# importing openpyxl module 
import openpyxl as xl

sns = 

base_path = 'C:/Users/Jdelpozo/Documents/prueba'

for sn in sns:
    for (root,dirs,files) in os.walk(base_path, topdown=True):
        for file in files:
            if file[-5:] == '.xlsx':
                # opening the source excel file 
                wb1 = xl.load_workbook(root + file) 
                ws1 = wb1.worksheets[14]
                #print('This is 1',ws1)
                # opening the destination excel file  
                filename1 ="C:/Users/Jdelpozo/Documents/Pressure_profiles_old.xlsx"
                wb2 = xl.load_workbook(filename1)
                newsheet = wb2.create_sheet(str(sn))
                # print('This is 2',ws2)

                # number of rows and columns in source excel file 
                min_r = 1
                min_c = 1

                mr = ws1.max_row 
                mc = ws1.max_column 
                
                # copying the cell values from source  
                # excel file to destination excel file 
                for i in range (1, mr + 1): 
                    for j in range (1, mc + 1): 
                        # reading cell value from source excel file 
                        c = ws1.cell(row = i, column = j) 
                
                        # writing the read value to destination excel file 
                        newsheet.cell(row = i, column = j).value = c.value 
                
                # saving the destination excel file 
                wb2.save(str(filename1))
            else:
                pass